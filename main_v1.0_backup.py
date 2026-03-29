# -*- coding: utf-8 -*-
import os
import sys
import json
import socket
import threading
import time
import datetime
import traceback


# ── Crash logger ──────────────────────────────────────────────────────────────
def _on_crash(exc_type, exc_value, exc_tb):
    try:
        from kivy.app import App
        app = App.get_running_app()
        p = os.path.join(app.user_data_dir if app else '.', 'crash.txt')
        with open(p, 'w') as f:
            traceback.print_exception(exc_type, exc_value, exc_tb, file=f)
    except Exception:
        pass
    sys.__excepthook__(exc_type, exc_value, exc_tb)

sys.excepthook = _on_crash

# ── Kivy imports ──────────────────────────────────────────────────────────────
from kivy.app import App
from kivy.clock import Clock
from kivy.metrics import dp, sp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.gridlayout import GridLayout
from kivy.uix.image import Image
from kivy.uix.label import Label
from kivy.uix.screenmanager import NoTransition, Screen, ScreenManager
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.uix.togglebutton import ToggleButton
from kivy.uix.widget import Widget
from kivy.utils import get_color_from_hex as hex_c
from kivy.graphics import Color, Rectangle

# ── Colors ────────────────────────────────────────────────────────────────────
NAV_BG   = hex_c('#1a2332')
C_BLUE   = hex_c('#3b82f6')
C_ORANGE = hex_c('#f97316')
C_RED    = hex_c('#ef4444')
C_GREEN  = hex_c('#22c55e')
C_DARK   = hex_c('#0f172a')
C_CYAN   = hex_c('#38bdf8')
C_TEXT   = hex_c('#1e293b')
C_MUTED  = hex_c('#64748b')
C_WHITE  = hex_c('#ffffff')
C_LIGHT  = hex_c('#ffffff')
C_ROW1   = hex_c('#f8fafc')

# ── Config ────────────────────────────────────────────────────────────────────
DEFAULT_CFG = {"ip": "192.168.1.100", "port": 8000, "mode": "manual", "interval": 2.0}

def cfg_path():
    return os.path.join(App.get_running_app().user_data_dir, 'config.json')

def load_cfg():
    try:
        with open(cfg_path(), 'r') as f:
            d = json.load(f)
        for k, v in DEFAULT_CFG.items():
            d.setdefault(k, v)
        return d
    except Exception:
        return dict(DEFAULT_CFG)

def save_cfg(cfg):
    with open(cfg_path(), 'w') as f:
        json.dump(cfg, f)

# ── History ───────────────────────────────────────────────────────────────────
def history_path():
    return os.path.join(App.get_running_app().user_data_dir, 'history.json')

def save_reading(weight, kind):
    if kind == 'nocomm':
        return
    try:
        recs = load_all_readings()
        recs.append({'ts': now_str(), 'weight': weight, 'kind': kind})
        if len(recs) > 50000:
            recs = recs[-50000:]
        with open(history_path(), 'w') as f:
            json.dump(recs, f)
    except Exception:
        pass

def load_all_readings():
    try:
        with open(history_path(), 'r') as f:
            return json.load(f)
    except Exception:
        return []

def filter_readings(recs, date_str=''):
    s = date_str.strip()
    if not s:
        return recs
    return [r for r in recs if r.get('ts', '').startswith(s)]

def export_excel(recs, path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Weight Readings'
        hf = Font(bold=True, color='FFFFFF')
        hfill = PatternFill('solid', fgColor='1A2332')
        for col, h in enumerate(['Date', 'Time', 'Weight', 'Status'], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill
            c.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 13
        ws.column_dimensions['B'].width = 11
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 10
        for rec in recs:
            ts = rec.get('ts', '')
            parts = ts.split('  ')
            date = parts[0] if parts else ts
            t    = parts[1] if len(parts) > 1 else ''
            ws.append([date, t, rec.get('weight', ''), rec.get('kind', '')])
        wb.save(path)
        return True
    except Exception as e:
        return str(e)

def share_file_android(path):
    try:
        from kivy.utils import platform
        if platform != 'android':
            return 'File saved: ' + path
        from jnius import autoclass
        from android import activity
        Intent    = autoclass('android.content.Intent')
        Uri       = autoclass('android.net.Uri')
        File      = autoclass('java.io.File')
        f = File(path)
        try:
            FileProvider = autoclass('androidx.core.content.FileProvider')
            ctx = activity.getApplicationContext()
            uri = FileProvider.getUriForFile(
                ctx, ctx.getPackageName() + '.fileprovider', f)
        except Exception:
            uri = Uri.fromFile(f)
        intent = Intent(Intent.ACTION_SEND)
        intent.setType('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        intent.putExtra(Intent.EXTRA_STREAM, uri)
        intent.putExtra(Intent.EXTRA_SUBJECT, 'Weight Readings Export')
        intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
        activity.startActivity(Intent.createChooser(intent, 'Share Excel'))
        return None
    except Exception as e:
        return str(e)

# ── Scale TCP ─────────────────────────────────────────────────────────────────
def scale_cmd(ip, port, cmd, timeout=5):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(timeout)
        s.connect((ip, int(port)))
        s.sendall(cmd.encode('ascii'))
        if cmd == 'W':
            buf = b''
            deadline = time.time() + timeout
            while time.time() < deadline:
                try:
                    chunk = s.recv(16)
                except socket.timeout:
                    break
                if not chunk:
                    break
                buf += chunk
                if len(buf.rstrip(b'\r\n')) >= 7:
                    break
            return buf.decode('ascii', errors='replace').strip()
        try:
            s.settimeout(1); s.recv(16)
        except Exception:
            pass
        return 'ok'

# ── Weight parsing ────────────────────────────────────────────────────────────
def now_str():
    return datetime.datetime.now().strftime('%d/%m/%Y  %H:%M:%S')

def parse_weight(raw):
    v = raw.strip().upper()
    if not v:
        return 'No Comm', 'nocomm'
    if v == 'H' or v == 'STOP':
        return 'OVER', 'over'
    if v == 'U' or v == 'UNDER' or v == 'UNDRE':
        return 'UNDER', 'under'
    return raw.strip(), 'normal'

# ── UI helpers ────────────────────────────────────────────────────────────────
class BgWidget(Widget):
    def __init__(self, color, **kwargs):
        super().__init__(**kwargs)
        self._bg_color = color
        self.bind(pos=self._redraw, size=self._redraw)
    def _redraw(self, *_):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self._bg_color); Rectangle(pos=self.pos, size=self.size)

class BgBox(BoxLayout):
    def __init__(self, color, **kwargs):
        super().__init__(**kwargs)
        self._bg_color = color
        self.bind(pos=self._redraw, size=self._redraw)
    def _redraw(self, *_):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self._bg_color); Rectangle(pos=self.pos, size=self.size)

def make_btn(text, bg, cb, **kwargs):
    kwargs.setdefault('size_hint', (1, None))
    if 'height' not in kwargs:
        kwargs['height'] = dp(54)
    btn = Button(text=text, background_normal='', background_color=bg,
                 color=C_WHITE, font_size=sp(16), bold=True, **kwargs)
    btn.bind(on_press=cb)
    return btn

def make_lbl(text, size=None, color=None, bold=False, halign='right', height=None):
    lbl = Label(text=text,
                font_size=size if size is not None else sp(14),
                color=color if color is not None else C_TEXT,
                bold=bold, halign=halign, size_hint_y=None,
                height=height if height is not None else dp(30))
    lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
    return lbl

def make_inp(hint, text='', input_filter=None):
    return TextInput(hint_text=hint, text=str(text), size_hint=(1, None),
                     height=dp(48), font_size=sp(16), multiline=False,
                     input_filter=input_filter, background_color=C_WHITE,
                     foreground_color=C_TEXT, padding=[dp(12), dp(10)],
                     write_tab=False)

# ── NavBar ────────────────────────────────────────────────────────────────────
class NavBar(BgBox):
    def __init__(self, sm, **kwargs):
        super().__init__(NAV_BG, orientation='horizontal',
                         size_hint=(1, None), height=dp(56), **kwargs)
        self.sm = sm
        self.add_widget(Widget())

        self.btn_history = Button(
            text='History', size_hint=(None, 1), width=dp(100),
            background_normal='', background_color=NAV_BG,
            color=hex_c('#94a3b8'), font_size=sp(13))
        self.btn_settings = Button(
            text='Settings', size_hint=(None, 1), width=dp(100),
            background_normal='', background_color=NAV_BG,
            color=hex_c('#94a3b8'), font_size=sp(13))
        self.btn_weight = Button(
            text='Weight', size_hint=(None, 1), width=dp(100),
            background_normal='', background_color=C_BLUE,
            color=C_WHITE, font_size=sp(13), bold=True)

        self.btn_weight.bind(on_press=lambda *_: self._go('weight'))
        self.btn_settings.bind(on_press=lambda *_: self._go('settings'))
        self.btn_history.bind(on_press=lambda *_: self._go('history'))
        self.add_widget(self.btn_history)
        self.add_widget(self.btn_settings)
        self.add_widget(self.btn_weight)

    def _go(self, name):
        self.sm.current = name
        btns = {'weight': self.btn_weight,
                'settings': self.btn_settings,
                'history': self.btn_history}
        for n, b in btns.items():
            if n == name:
                b.background_color = C_BLUE; b.color = C_WHITE
            else:
                b.background_color = NAV_BG; b.color = hex_c('#94a3b8')

# ── Weight Screen ─────────────────────────────────────────────────────────────
class WeightScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._auto_event = None
        self._busy = False
        self._build_ui()

    def _build_ui(self):
        root = BoxLayout(orientation='vertical', padding=dp(16), spacing=dp(12))

        logo = Image(source='logo.jpg', size_hint=(1, None), height=dp(120),
                     allow_stretch=True, keep_ratio=True)
        root.add_widget(logo)

        card = BgBox(C_DARK, orientation='vertical', size_hint=(1, None),
                     height=dp(190), padding=dp(20), spacing=dp(4))

        card.add_widget(make_lbl('Current Weight', size=sp(12),
                                  color=hex_c('#94a3b8'), halign='center', height=dp(20)))

        self.lbl_weight = Label(text='---', font_size=sp(68), bold=True,
                                color=C_CYAN, size_hint_y=None, height=dp(90),
                                halign='center')
        self.lbl_weight.bind(
            size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
        card.add_widget(self.lbl_weight)

        card.add_widget(make_lbl('kg', size=sp(14), color=hex_c('#475569'),
                                  halign='center', height=dp(18)))

        self.lbl_status = make_lbl('Ready', size=sp(12),
                                    color=hex_c('#64748b'), halign='center', height=dp(20))
        card.add_widget(self.lbl_status)
        root.add_widget(card)

        btn_grid = GridLayout(cols=1, spacing=dp(10), size_hint=(1, None))
        btn_grid.bind(minimum_height=btn_grid.setter('height'))

        self.btn_get = make_btn('Weight', C_BLUE, self._on_get)
        btn_grid.add_widget(self.btn_get)

        row2 = BoxLayout(orientation='horizontal', spacing=dp(10),
                          size_hint=(1, None), height=dp(54))
        row2.add_widget(make_btn('Tare', C_ORANGE, self._on_tare))
        row2.add_widget(make_btn('Zero', C_RED, self._on_zero))
        btn_grid.add_widget(row2)
        root.add_widget(btn_grid)

        root.add_widget(Widget())

        footer = BoxLayout(orientation='vertical', size_hint=(1, None),
                           height=dp(72), spacing=dp(1))
        for line in ['Ver: 1.0', 'Shekel Scale 2008 LTD',
                     'Tel: 04-6629100', 'Itzik Zrihan']:
            footer.add_widget(make_lbl(line, size=sp(10),
                                       color=hex_c('#94a3b8'), halign='center', height=dp(16)))
        root.add_widget(footer)
        self.add_widget(root)

    def on_enter(self):
        cfg = load_cfg()
        if cfg.get('mode') == 'auto':
            self._start_auto(cfg)
        else:
            self._stop_auto(); self._set_btn_normal()

    def on_leave(self):
        self._stop_auto()

    def _on_get(self, *_):
        if not self._busy:
            threading.Thread(target=self._fetch, daemon=True).start()

    def _fetch(self):
        if self._busy:
            return
        self._busy = True
        try:
            cfg = load_cfg()
            val = scale_cmd(cfg['ip'], cfg['port'], 'W')
            display, kind = parse_weight(val)
            ts = now_str()
            Clock.schedule_once(lambda *_, d=display, k=kind, t=ts: self._set_weight(d, k, t))
        except Exception:
            ts = now_str()
            Clock.schedule_once(lambda *_, t=ts: self._set_weight('No Comm', 'nocomm', t))
        finally:
            self._busy = False

    def _on_tare(self, *_):
        threading.Thread(target=self._send_cmd, args=('T', 'Tare done'), daemon=True).start()

    def _on_zero(self, *_):
        threading.Thread(target=self._send_cmd, args=('Z', 'Zero done'), daemon=True).start()

    def _send_cmd(self, cmd, msg):
        try:
            cfg = load_cfg()
            scale_cmd(cfg['ip'], cfg['port'], cmd)
            Clock.schedule_once(lambda *_: self._set_status(msg, hex_c('#86efac')))
        except Exception as e:
            Clock.schedule_once(lambda *_: self._set_status(str(e)[:40], hex_c('#fca5a5')))

    def _start_auto(self, cfg):
        self._stop_auto()
        self.btn_get.text = 'Auto ON'
        self.btn_get.background_color = hex_c('#475569')
        self.btn_get.disabled = True
        interval = max(0.5, float(cfg.get('interval', 2.0)))
        self._on_get()
        self._auto_event = Clock.schedule_interval(lambda *_: self._on_get(), interval)

    def _stop_auto(self):
        if self._auto_event:
            self._auto_event.cancel(); self._auto_event = None

    def _set_btn_normal(self):
        self.btn_get.text = 'Weight'
        self.btn_get.background_color = C_BLUE
        self.btn_get.disabled = False

    def _set_weight(self, val, kind='normal', ts=''):
        colors = {'normal': C_CYAN, 'over': hex_c('#f97316'),
                  'under': hex_c('#f59e0b'), 'nocomm': hex_c('#f87171')}
        sizes  = {'normal': sp(62), 'over': sp(52),
                  'under': sp(52), 'nocomm': sp(40)}
        self.lbl_weight.text      = val
        self.lbl_weight.color     = colors.get(kind, C_CYAN)
        self.lbl_weight.font_size = sizes.get(kind, sp(52))
        self.lbl_status.text      = ts
        self.lbl_status.color     = hex_c('#64748b')
        save_reading(val, kind)

    def _set_status(self, msg, color=None):
        self.lbl_status.text = msg
        if color is not None:
            self.lbl_status.color = color

# ── Settings Screen ───────────────────────────────────────────────────────────
class SettingsScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._build_ui()

    def _build_ui(self):
        root = BoxLayout(orientation='vertical', padding=dp(16), spacing=dp(10))
        root.add_widget(make_lbl('Connection Settings', size=sp(18), bold=True))

        root.add_widget(make_lbl('IP Address'))
        self.inp_ip = make_inp('192.168.1.100')
        root.add_widget(self.inp_ip)

        root.add_widget(make_lbl('Port'))
        self.inp_port = make_inp('8000', input_filter='int')
        root.add_widget(self.inp_port)

        root.add_widget(make_lbl('Mode'))
        mode_row = BoxLayout(orientation='horizontal', size_hint=(1, None),
                              height=dp(48), spacing=dp(8))
        self.btn_manual = ToggleButton(
            text='Manual', group='mode', state='down',
            background_normal='', background_down='',
            background_color=C_BLUE, color=C_WHITE, font_size=sp(15), bold=True)
        self.btn_auto = ToggleButton(
            text='Auto', group='mode',
            background_normal='', background_down='',
            background_color=hex_c('#e2e8f0'), color=C_TEXT, font_size=sp(15))
        self.btn_manual.bind(on_press=lambda *_: self._mode('manual'))
        self.btn_auto.bind(on_press=lambda *_: self._mode('auto'))
        mode_row.add_widget(self.btn_auto)
        mode_row.add_widget(self.btn_manual)
        root.add_widget(mode_row)

        self.lbl_interval = make_lbl('Interval (seconds)')
        root.add_widget(self.lbl_interval)
        self.inp_interval = make_inp('2', input_filter='float')
        root.add_widget(self.inp_interval)

        root.add_widget(Widget(size_hint_y=None, height=dp(8)))

        btn_row = BoxLayout(orientation='horizontal', size_hint=(1, None),
                             height=dp(54), spacing=dp(10))
        btn_row.add_widget(make_btn('Save', C_GREEN, self._save, size_hint=(1, 1)))
        btn_row.add_widget(make_btn('Test', C_BLUE, self._test, size_hint=(1, 1)))
        root.add_widget(btn_row)

        self.lbl_result = make_lbl('', size=sp(13), color=C_MUTED)
        root.add_widget(self.lbl_result)
        root.add_widget(Widget())
        self.add_widget(root)

    def on_enter(self):
        cfg = load_cfg()
        self.inp_ip.text       = str(cfg.get('ip', ''))
        self.inp_port.text     = str(cfg.get('port', ''))
        self.inp_interval.text = str(cfg.get('interval', 2.0))
        if cfg.get('mode') == 'auto':
            self.btn_auto.state = 'down'; self.btn_manual.state = 'normal'
            self._mode('auto')
        else:
            self.btn_manual.state = 'down'; self.btn_auto.state = 'normal'
            self._mode('manual')

    def _mode(self, mode):
        if mode == 'auto':
            self.btn_auto.background_color = C_BLUE;   self.btn_auto.color   = C_WHITE
            self.btn_manual.background_color = hex_c('#e2e8f0'); self.btn_manual.color = C_TEXT
            self.inp_interval.disabled = False
        else:
            self.btn_manual.background_color = C_BLUE; self.btn_manual.color = C_WHITE
            self.btn_auto.background_color = hex_c('#e2e8f0');   self.btn_auto.color   = C_TEXT
            self.inp_interval.disabled = True

    def _save(self, *_):
        ip = self.inp_ip.text.strip()
        if not ip:
            self._result('Enter IP address', hex_c('#ef4444')); return
        try:
            port = int(self.inp_port.text.strip()); assert 1 <= port <= 65535
        except Exception:
            self._result('Invalid port', hex_c('#ef4444')); return
        try:
            interval = float(self.inp_interval.text.strip()); assert interval > 0
        except Exception:
            self._result('Invalid interval', hex_c('#ef4444')); return
        mode = 'auto' if self.btn_auto.state == 'down' else 'manual'
        try:
            save_cfg({'ip': ip, 'port': port, 'mode': mode, 'interval': interval})
            App.get_running_app().nav._go('weight')
        except Exception as e:
            self._result('Save error: ' + str(e)[:30], hex_c('#ef4444'))

    def _test(self, *_):
        self._result('Testing...', hex_c('#60a5fa'))
        ip = self.inp_ip.text.strip()
        try:
            port = int(self.inp_port.text.strip())
        except Exception:
            self._result('Invalid port', hex_c('#ef4444')); return
        def _do():
            try:
                val = scale_cmd(ip, port, 'W', timeout=4)
                Clock.schedule_once(lambda *_: self._result('Connected: ' + val, hex_c('#22c55e')))
            except Exception as e:
                Clock.schedule_once(lambda *_: self._result(str(e)[:40], hex_c('#ef4444')))
        threading.Thread(target=_do, daemon=True).start()

    def _result(self, msg, color=None):
        self.lbl_result.text = msg
        if color is not None: self.lbl_result.color = color

# ── History Screen ────────────────────────────────────────────────────────────
class HistoryScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._all_recs = []
        self._last_excel = None
        self._build_ui()

    def _build_ui(self):
        root = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))

        root.add_widget(make_lbl('Weight History', size=sp(18), bold=True))

        # Filter row
        fr = BoxLayout(orientation='horizontal', size_hint=(1, None),
                       height=dp(44), spacing=dp(6))
        self.inp_date = TextInput(
            hint_text='DD/MM/YYYY', size_hint=(1, 1),
            font_size=sp(14), multiline=False,
            background_color=C_WHITE, foreground_color=C_TEXT,
            padding=[dp(10), dp(12)])
        btn_today = Button(text='Today', size_hint=(None, 1), width=dp(78),
                           background_normal='', background_color=C_BLUE,
                           color=C_WHITE, font_size=sp(13), bold=True)
        btn_all   = Button(text='All', size_hint=(None, 1), width=dp(55),
                           background_normal='', background_color=hex_c('#475569'),
                           color=C_WHITE, font_size=sp(13))
        btn_today.bind(on_press=self._filter_today)
        btn_all.bind(on_press=self._filter_all)
        fr.add_widget(self.inp_date); fr.add_widget(btn_today); fr.add_widget(btn_all)
        root.add_widget(fr)

        self.lbl_count = make_lbl('', size=sp(11), color=C_MUTED, height=dp(18))
        root.add_widget(self.lbl_count)

        # Header
        root.add_widget(self._row('Date', 'Time', 'Weight',
                                   bold=True, bg=hex_c('#1a2332'), fg=C_WHITE))

        # List
        scroll = ScrollView(size_hint=(1, 1))
        self.list_box = GridLayout(cols=1, size_hint_y=None, spacing=dp(1))
        self.list_box.bind(minimum_height=self.list_box.setter('height'))
        scroll.add_widget(self.list_box)
        root.add_widget(scroll)

        # Action buttons
        ab = BoxLayout(orientation='horizontal', size_hint=(1, None),
                       height=dp(52), spacing=dp(10))
        ab.add_widget(make_btn('Export Excel', C_GREEN, self._export, size_hint=(1, 1)))
        ab.add_widget(make_btn('Send Email',   C_BLUE,  self._email,  size_hint=(1, 1)))
        root.add_widget(ab)

        self.lbl_msg = make_lbl('', size=sp(11), color=C_MUTED, height=dp(20))
        root.add_widget(self.lbl_msg)
        self.add_widget(root)

    def _row(self, date, time_val, weight, bold=False, bg=None, fg=None):
        row = BgBox(bg or C_WHITE, orientation='horizontal',
                    size_hint=(1, None), height=dp(30), padding=[dp(4), 0])
        for text, hint in [(date, 0.38), (time_val, 0.32), (weight, 0.30)]:
            lbl = Label(text=text, size_hint=(hint, 1),
                        font_size=sp(11), bold=bold,
                        color=fg or C_TEXT, halign='left')
            lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
            row.add_widget(lbl)
        return row

    def on_enter(self):
        self._load(self.inp_date.text.strip())

    def _load(self, date_filter=''):
        self._all_recs = load_all_readings()
        recs = filter_readings(self._all_recs, date_filter)
        shown = list(reversed(recs[-300:]))
        self.list_box.clear_widgets()
        for i, r in enumerate(shown):
            ts = r.get('ts', '')
            parts = ts.split('  ')
            date = parts[0] if parts else ts
            t    = parts[1] if len(parts) > 1 else ''
            bg = C_ROW1 if i % 2 == 0 else C_WHITE
            self.list_box.add_widget(self._row(date, t, r.get('weight', ''), bg=bg))
        self.lbl_count.text = (f'Showing {len(shown)} of {len(self._all_recs)} records')

    def _filter_today(self, *_):
        today = datetime.datetime.now().strftime('%d/%m/%Y')
        self.inp_date.text = today
        self._load(today)

    def _filter_all(self, *_):
        self.inp_date.text = ''
        self._load('')

    def _export(self, *_):
        recs = filter_readings(self._all_recs, self.inp_date.text.strip())
        if not recs:
            self._msg('No records', hex_c('#ef4444')); return
        path = os.path.join(App.get_running_app().user_data_dir, 'weight_export.xlsx')
        result = export_excel(recs, path)
        if result is True:
            self._last_excel = path
            self._msg(f'Excel saved ({len(recs)} rows)', hex_c('#22c55e'))
        else:
            self._msg('Error: ' + str(result)[:40], hex_c('#ef4444'))

    def _email(self, *_):
        if not self._last_excel:
            self._export()
        if not self._last_excel:
            return
        result = share_file_android(self._last_excel)
        if result:
            self._msg(result[:50], hex_c('#ef4444'))

    def _msg(self, text, color=None):
        self.lbl_msg.text = text
        if color is not None: self.lbl_msg.color = color

# ── App ───────────────────────────────────────────────────────────────────────
class ScaleApp(App):
    def build(self):
        from kivy.core.window import Window
        Window.clearcolor = C_LIGHT

        sm = ScreenManager(transition=NoTransition())
        sm.add_widget(WeightScreen(name='weight'))
        sm.add_widget(SettingsScreen(name='settings'))
        sm.add_widget(HistoryScreen(name='history'))

        self.nav = NavBar(sm)
        root = BoxLayout(orientation='vertical')
        root.add_widget(self.nav)
        root.add_widget(sm)
        return root

if __name__ == '__main__':
    ScaleApp().run()
